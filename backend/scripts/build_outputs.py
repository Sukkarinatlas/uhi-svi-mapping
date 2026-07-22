# -*- coding: utf-8 -*-
"""Assemble master table and build Excel + Population_Density GeoTIFF (1km, EPSG:4326)."""
import json, os
import numpy as np
import geopandas as gpd
import pandas as pd
import rasterio
from rasterio import features
from rasterio.transform import from_origin

OUTDIR = r"e:\Gistda_GeoAI_VibeCode\GISTDA GeoAI & Vibe Coding - Material-20260713T022955Z-2-001\GISTDA GeoAI & Vibe Coding - Material\UHI_WEB\output"
os.makedirs(OUTDIR, exist_ok=True)
SP = os.path.dirname(os.path.abspath(__file__))

# ---- population (validated, sums to 1,645,985) ----
pop = {r["code6"]: r for r in json.load(open(os.path.join(SP,"admin_pop_final.json"),encoding="utf-8"))}

# ---- Thai names (kongvut) ----
tb  = {str(x["id"]): x for x in json.load(open(os.path.join(SP,"kv_tambon.json"),encoding="utf-8"))}
dis = {str(x["id"]): x for x in json.load(open(os.path.join(SP,"kv_district.json"),encoding="utf-8"))}

# ---- geometry + area (HDX COD adm3) ----
g = gpd.read_file(os.path.join(SP,"tha_shp","tha_admin3.shp"))
ch = g[g["adm1_pcode"]=="TH20"].copy()
ch["code6"] = ch["adm3_pcode"].str[2:]
ch["code4"] = ch["code6"].str[:4]

def th_tambon(c6): return tb.get(str(int(c6)),{}).get("name_th","")
def en_tambon(c6): return tb.get(str(int(c6)),{}).get("name_en","")
def th_amphoe(c4): return dis.get(str(int(c4)),{}).get("name_th","")
def en_amphoe(c4): return dis.get(str(int(c4)),{}).get("name_en","")

ch["tambon_th"]   = ch["code6"].map(th_tambon)
ch["tambon_en"]   = ch["code6"].map(en_tambon)
ch["amphoe_th"]   = ch["code4"].map(th_amphoe)
ch["amphoe_en"]   = ch["code4"].map(en_amphoe)
ch["male"]        = ch["code6"].map(lambda c: pop.get(c,{}).get("male",0)).astype(int)
ch["female"]      = ch["code6"].map(lambda c: pop.get(c,{}).get("female",0)).astype(int)
ch["total_pop"]   = ch["code6"].map(lambda c: pop.get(c,{}).get("total",0)).astype(int)
ch["area_km2"]    = ch["area_sqkm"].astype(float).round(4)
ch["pop_density"] = (ch["total_pop"]/ch["area_km2"]).round(2)   # persons per km^2

assert ch["total_pop"].sum()==1645985, ch["total_pop"].sum()
print("Master table rows:", len(ch), "total pop:", ch["total_pop"].sum())

# ---- save GeoPackage + GeoJSON (bonus, same attributes) ----
keep=["code6","adm3_pcode","province_th","province_en","amphoe_th","amphoe_en",
      "tambon_th","tambon_en","male","female","total_pop","area_km2","pop_density","geometry"]
ch["province_th"]="ชลบุรี"; ch["province_en"]="Chon Buri"
gout=ch[keep].copy()
gout.to_file(os.path.join(OUTDIR,"Chonburi_Population_Tambon_Dec2568.gpkg"), driver="GPKG")
gout.to_file(os.path.join(OUTDIR,"Chonburi_Population_Tambon_Dec2568.geojson"), driver="GeoJSON")
print("wrote GeoPackage + GeoJSON")

# ---- Excel ----
df = pd.DataFrame(ch.drop(columns="geometry"))
df = df[["code6","amphoe_th","tambon_th","tambon_en","male","female","total_pop","area_km2","pop_density"]]
df = df.sort_values("code6").reset_index(drop=True)
df.columns = ["รหัสตำบล(6หลัก)","อำเภอ","ตำบล","Tambon (EN)","ชาย","หญิง",
              "ประชากรรวม","พื้นที่ (ตร.กม.)","ความหนาแน่น (คน/ตร.กม.)"]
xlsx = os.path.join(OUTDIR,"Chonburi_Population_Tambon_Dec2568.xlsx")
with pd.ExcelWriter(xlsx, engine="openpyxl") as xw:
    df.to_excel(xw, index=False, sheet_name="ประชากรรายตำบล")
    # summary by amphoe
    amp = df.groupby("อำเภอ", sort=False).agg(
        ตำบล=("ตำบล","count"), ชาย=("ชาย","sum"), หญิง=("หญิง","sum"),
        ประชากรรวม=("ประชากรรวม","sum")).reset_index()
    amp.to_excel(xw, index=False, sheet_name="สรุปรายอำเภอ")
# formatting
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
wb=load_workbook(xlsx)
hdrfill=PatternFill("solid",fgColor="4472C4"); hdrfont=Font(bold=True,color="FFFFFF")
thin=Side(style="thin",color="D9D9D9"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
for ws in wb.worksheets:
    for c in ws[1]:
        c.fill=hdrfill; c.font=hdrfont; c.alignment=Alignment(horizontal="center",vertical="center")
    for col in ws.columns:
        w=max(len(str(c.value)) for c in col if c.value is not None)
        ws.column_dimensions[col[0].column_letter].width=min(max(w+3,10),16)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.border=border
            if isinstance(c.value,(int,float)): c.number_format="#,##0.##"
    ws.freeze_panes="A2"
# total row on sheet1
ws=wb["ประชากรรายตำบล"]; last=ws.max_row+1
ws.cell(last,1,"รวมทั้งจังหวัด"); ws.cell(last,5,int(df["ชาย"].sum())); ws.cell(last,6,int(df["หญิง"].sum()))
ws.cell(last,7,int(df["ประชากรรวม"].sum())); ws.cell(last,8,round(df["พื้นที่ (ตร.กม.)"].sum(),2))
for cc in ws[last]:
    cc.font=Font(bold=True); cc.fill=PatternFill("solid",fgColor="D9E1F2")
    if isinstance(cc.value,(int,float)): cc.number_format="#,##0.##"
wb.save(xlsx)
print("wrote Excel:", xlsx)

# ---- GeoTIFF: population density, ~1km grid, EPSG:4326 ----
# 1 km expressed in degrees at WGS84 (~0.008983 deg); density is intensive so value is
# independent of exact pixel size. Grid aligned to whole-degree origin.
RES = 1000.0/111320.0   # ~0.0089832 deg  (~1 km)
gj = gout.to_crs(4326)
minx,miny,maxx,maxy = gj.total_bounds
# snap to grid
minx=np.floor(minx/RES)*RES; miny=np.floor(miny/RES)*RES
maxx=np.ceil(maxx/RES)*RES;  maxy=np.ceil(maxy/RES)*RES
W=int(round((maxx-minx)/RES)); H=int(round((maxy-miny)/RES))
transform=from_origin(minx, maxy, RES, RES)
shapes=[(geom, val) for geom,val in zip(gj.geometry, gj["pop_density"])]
NODATA=-9999.0
arr=features.rasterize(shapes, out_shape=(H,W), transform=transform,
                       fill=NODATA, all_touched=False, dtype="float32")
tif=os.path.join(OUTDIR,"Chonburi_Population_Density_Tambon_1km_EPSG4326.tif")
with rasterio.open(tif,"w",driver="GTiff",height=H,width=W,count=1,dtype="float32",
                   crs="EPSG:4326",transform=transform,nodata=NODATA,
                   compress="LZW",tiled=True) as dst:
    dst.write(arr,1)
    dst.update_tags(1, UNIT="persons_per_km2", SOURCE="DOPA statMONTH yymm=6812 (Dec 2568)",
                    BOUNDARY="HDX/RTSD COD adm3", NOTE="Value=tambon registered pop / tambon area(km2)")
    dst.set_band_description(1,"Population density (persons/km2), Dec 2568")
valid=arr[arr!=NODATA]
print(f"wrote GeoTIFF: {tif}")
print(f"  size {W}x{H} px, res~{RES:.6f} deg (~1km), density min={valid.min():.1f} max={valid.max():.1f} mean={valid.mean():.1f}")
